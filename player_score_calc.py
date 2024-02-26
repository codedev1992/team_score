import streamlit as st 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from io import BytesIO
from openpyxl.styles import Alignment, Font
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill
import math
import random

DEBUG = True

teams_file = st.file_uploader("upload teams excel file.")

st.info("NOTE: For generating My Teams, the player value should start from A19")
c1,c2 = st.columns(2)
with c1:
    process_button = st.button("process")
with c2:
    ic1,ic2,ic3 = st.columns(3)
    with ic1:
        use_user_weight = st.checkbox("Use Manual Weight", key="use_manual_weight")
    with ic2:
        is_retro_game = st.checkbox("Is Retro Game", key="is_retro_game")
    with ic3: 
        my_team_formation = st.button("My Team Process")
        

players_sheet_name  = "PlayersList"
sheet_to_use = "Copy of Teams"
teams_list_sheet_name = "TeamsList"
my_team_sheet_name = "My Teams"
sel_value_sheet_name = "Sel Value"


C_COLOR = ["FF00FF00","FF00B050"]
VC_COLOR = ["FFFFFF00"] 
RED = "FFFF0000"
BLACK = 'FF000000'
TEAMS = []


PLAYER_TYPE_RULES = {
"W": {"min" : 1, "max": 4},
"Ba" :{"min" : 3, "max": 6},
"A": {"min" : 1, "max": 4},
"Bo": {"min" : 3, "max": 6},
}



def util_transpose(l1):
    l2 = [[row[i] for row in l1] for i in range(len(l1[0]))]
    return l2

def util_T_to_1D_list(_2dT):
    result = []

    # Iterate over the range of the length of the sublists
    for i in range(len(_2dT)):
        # Rotate the sublist by one position to the left and add the first element to the result list
        sublist = _2dT[i]
        # Rotate by taking elements from 1 to end and then the first element
        rotate_idx = i%len(sublist)

        if i !=0:
            emlt_to_rotte = sublist[rotate_idx]
            sublist.pop(rotate_idx)
            sublist.append(emlt_to_rotte)
        rotated_sublist =  sublist #sublist[i:] + sublist[:i]
    
        # Append the first element of the rotated sublist to the result

        for e in rotated_sublist:
            if e is not None:
                result.append(e)
    
    return result


def get_team_status(my_team_players,my_team):
    teams_status = []
 
    for update_team in my_team:
        tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
        for pname in update_team:
            pyr = my_team_players.get(pname,{})
            ptype = pyr.get("type","W")
            pcolor = pyr.get("color","b")
            if ptype is not None:
                tems_cnt[ptype] = tems_cnt[ptype] + 1
            
            if pcolor is not None:
                tems_cnt[pcolor] = tems_cnt[pcolor] + 1
    
        teams_status.append(tems_cnt)
    
    return teams_status
 

def check_and_replace_players(my_team_players,my_team,team_comb_dict,min_max_rules):
    all_team_current_status = get_team_status(my_team_players,my_team)

    for idx,team in enumerate(my_team):
        if len(team) < 11:
            # print(idx)
            no_plyr_required = 11 - len(team)
            team_status = all_team_current_status[idx]
            #print(team_status,team_comb_dict)
            clr_to_replace = 'b' if team_status['b'] > team_status['r'] else 'r'
            # print(team_comb_dict)
            for pcat,pname_list in team_comb_dict.items():
                if no_plyr_required > 0:
                    if min_max_rules["max"][pcat] > team_status[pcat]:
                        if len(team_comb_dict[pcat]) > 0:
                            for pyrs_list in pname_list:
                                if pyrs_list:
                                    if my_team_players[pyrs_list[0]].get("color") == clr_to_replace:
                                        suitable_players_to_replace = []
                                        for pname,prop in  my_team_players.items():
                                            if prop["color"] == clr_to_replace and prop["type"] == pcat and pyrs_list[0] != pname:
                                                
                                                for mt in my_team:
                                                    if pname in mt and pyrs_list[0] not in mt:
                                                        idx_of_plr_to_be_replaced = mt.index(pname)
                                                        mt[idx_of_plr_to_be_replaced] = pyrs_list[0] 
                                                        team.append(pname)
                                                        # print(pname, pyrs_list[0])
                                                        suitable_players_to_replace.append(pname)
                                                        no_plyr_required = no_plyr_required - 1
                                                        break
                                            
                                            if len(team) == 11:
                                                break

                                    if len(team) == 11:
                                        break
                if len(team) == 11:
                    break
                                            

                            
def plyer_to_type(team,mapping):
    team_type_count = {"W":0, "Ba":0, "Bo": 0, "A":0}

    for pl in team:
        team_type_count[mapping[pl].get("type")] = team_type_count[mapping[pl].get("type")] + 1
    
    return team_type_count

def player_to_color(team,mapping):
    team_type_count = {"r":0, "b":0}

    for pl in team:
        team_type_count[mapping[pl].get("color")] = team_type_count[mapping[pl].get("color")] + 1
    
    return team_type_count

def can_add_this_player_by_color(team,mapping, new_player):
    pl_to_clr_map = player_to_color(team,mapping)

    pl_to_clr_map[mapping[new_player].get("color")] = pl_to_clr_map[mapping[new_player].get("color")] + 1

    r_clr, b_clr = pl_to_clr_map.get("r",0), pl_to_clr_map.get("b",0)
    
    min_v, max_v = (r_clr, b_clr) if r_clr <= b_clr else (b_clr, r_clr)

    if ((min_v + max_v) <= 11 and max_v <=7):
        return True 
    else:
        return False

def get_team_combination(no_team,player_to_type,player_weights):
    #print(player_to_type,player_weights)

    output = {
        "W" :[],
        "Ba":[],
        "Bo":[],
        "A":[]
    }

    total_possible_players = no_team * 12 
    w,ba,bo,a = 2,4,4,2

    w1,ba1,bo1,a1 = len(player_to_type["W"]),len(player_to_type["Ba"]),len(player_to_type["Bo"]),len(player_to_type["A"])

    exp_w, exp_ba, exp_bo, exp_a = w*no_team, ba*no_team, bo*no_team, a*no_team

    wgt_w, wgt_ba,wgt_bo,wgt_a = math.ceil(exp_w/w1),math.ceil(exp_ba/ba1),math.ceil(exp_bo/bo1),math.ceil(exp_a/a1)

    diff_weight = {
        "W":(wgt_w*w1)-exp_w, 
        "Ba":(wgt_ba*ba1)-exp_ba,
        "Bo":(wgt_bo*bo1)-exp_bo,
        "A":(wgt_a*a1)-exp_a
    }

    p_weights = {
        "W":[wgt_w]*w1,
        "Ba":[wgt_ba]*ba1,
        "Bo":[wgt_bo]*bo1,
        "A":[wgt_a]*a1
    }

     
    for k,v in diff_weight.items():
        v = diff_weight[k]
        if v > 0:
            for i in range(-v,0):
                new_va = p_weights[k][i] - 1
                p_weights[k][i]  = new_va

    p_weights_to_dict = p_weights

    for k,v in p_weights_to_dict.items():
        for idx in range(len(v)):
            pname = player_to_type[k][idx]
            if (player_weights.get(pname,None) is None or player_weights.get(pname) == -1):
                output[k].append([player_to_type[k][idx]]*v[idx])
            else: 
                output[k].append([player_to_type[k][idx]]*player_weights[pname])

    # if DEBUG:
    #     print(output)

 
    return output

def manual_generate_my_teams(master_wb, file_name):
    no_team = st.session_state["input_team_generation_count"]

    is_manual_weight = st.session_state["use_manual_weight"]

    player_credit = {}
    if is_manual_weight:
        
        for key in st.session_state:
        # Check if the key belongs to your input fields
            if key.startswith("input_myteam_"):
                # Access the value of each input field
                value = st.session_state[key]
                # Process the value as needed
                pname = re.sub("input_myteam_[\d]+_", "", key)

                #print(pname, value)
                if value is None or value == "":
                    player_credit[pname] = -1
                else:
                    player_credit[pname] = eval(value)

    is_weight_inputted = False 


    wb = master_wb #load_workbook(exel_file, read_only=False)
    sheet_names = wb.sheetnames
    if my_team_sheet_name in sheet_names:
        my_team_players = {}
        my_team_sheet = wb[my_team_sheet_name]
        gph_idx = f"A19:C45"
        
        rnk = 1 
        top_11_players = []
        bottom_11_player = []
        player_to_type = {
            "A":[],
            "Ba" :[],
            "Bo":[],
            "W":[]
        }

        for row in my_team_sheet[gph_idx]:
            if row[0] is not None and row[0].value is not None:
                
                if row[0].font and row[0].font.color:
                    font_color = row[0].font.color.rgb
                    if font_color  == RED:
                        color = "r"
                    else:
                        color = "b"
                else:
                    color = "b"
                                
                my_team_players[row[0].value] = {"color": color, "rank" :rnk, 
                                                 "type":  row[1].value,
                                                   "weight":player_credit.get(row[0].value, 0)}

                player_to_type[row[1].value].append(row[0].value)

                if rnk <=11:
                    top_11_players.append(row[0].value)
                else:
                    bottom_11_player.append(row[0].value)
                rnk = rnk + 1

        btm_players_2d_matrx = []
        player_btm_min, player_btm_max = 99999,0 
        for btm_player in bottom_11_player:
            wgt = my_team_players.get(btm_player,{}).get("weight", 0)
            if wgt > player_btm_max:
                player_btm_max = wgt 
            if wgt <= player_btm_min:
                player_btm_min = wgt

        for btm_player in bottom_11_player:
            wgt = my_team_players.get(btm_player,{}).get("weight", 0)
            btm_players_2d_matrx.append([btm_player]*wgt + [None]* (player_btm_max - wgt))
        
        btm_players_2d_T = util_transpose(btm_players_2d_matrx)

        my_team = []
        for i in range(no_team):
            plrs = []
            for p in top_11_players:
                plrs.append(p)
            my_team.append(plrs)

        #print(btm_players_2d_T)

        rslt = util_T_to_1D_list(btm_players_2d_T)

        pidx, idx,lidx = 0, 0, len(bottom_11_player)-1
        s_tm_idx = 0
        e_tm_idx = 0
        for j in range(len(bottom_11_player)-1, -1,-1):
             
            wgt = my_team_players[bottom_11_player[idx]].get("weight", 0)
            
            plyr_range = wgt
            # if wgt < len(rslt):
            #     plyr_range = wgt
            # else:
            #     plyr_range = len(rslt)

            # sub_list_players = rslt[0:plyr_range-1]

            # rslt = rslt[plyr_range:]

            e_tm_idx = s_tm_idx + plyr_range

            for m_tm_idx in range(s_tm_idx, e_tm_idx,1):
                if pidx < len(rslt):
                    if rslt[pidx] not in my_team[(m_tm_idx % no_team)]:
                        my_team[(m_tm_idx % no_team)][lidx] = rslt[pidx]
                        pidx = pidx + 1
                    else:
                        if (pidx + 1) < len(rslt):
                            tmp = rslt[pidx] 
                            rslt[pidx] = rslt[pidx + 1]
                            rslt[pidx + 1] = tmp 
                            my_team[(m_tm_idx % no_team)][lidx] = rslt[pidx]
                            pidx = pidx + 1



            s_tm_idx = e_tm_idx

            idx = idx + 1
            lidx = lidx - 1

        teams_status = []
        t_count = 1
        for update_team in my_team:
            tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
            for pname in update_team:
                pyr = my_team_players.get(pname,{})
                ptype = pyr.get("type","W")
                pcolor = pyr.get("color","b")
                if ptype is not None:
                    tems_cnt[ptype] = tems_cnt[ptype] + 1
                
                if pcolor is not None:
                    tems_cnt[pcolor] = tems_cnt[pcolor] + 1
        
            teams_status.append(tems_cnt)
            t_count = t_count + 1

        team_count = 1
        last_col_name = get_column_letter(no_team)

        team_number_range = f"A1:{last_col_name}1"
        min_col, min_row, max_col, max_row = range_boundaries(team_number_range)
        team_idx = 1
        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = team_idx
            team_idx = team_idx + 1

        p_count = 0
        for tm in my_team:
            p_count = p_count + len(tm)
            if len(tm) < 11:
                missing_no_of_players = 11 - len(tm)
                for i in range(missing_no_of_players):
                    tm.append("NOT ABLE TO FILL")
            

        #print(team_comb_dict)

        write_range = f"A2:{last_col_name}12"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        #print(min_col, min_row, max_col, max_row)
        for col in range(min_col, max_col + 1):
            max_pname_len = 0
            for row in range(min_row, max_row + 1):
                pname = my_team[col-1][row-2]
                cell = my_team_sheet.cell(row=row, column=col)
                if len(str(pname)) > max_pname_len:
                    max_pname_len = len(str(pname))

                if pname in bottom_11_player:
                    cell.fill = PatternFill(start_color="FBDAD7", fill_type='solid')
                f_color = my_team_players.get(pname,{}).get("color","b")
                if f_color == "r":
                    red_font = Font(color=RED) 
                    cell.font = red_font
                else:
                    black_font = Font(color=BLACK) 
                    cell.font = black_font

                cell.value = pname
            
            col_letter = get_column_letter(col)
            my_team_sheet.column_dimensions[col_letter].width = max_pname_len
        
        
        write_range = f"A13:{last_col_name}18"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        for col in range(min_col, max_col + 1):
            tm_status = teams_status[col-1]

            a_count = tm_status.get("A", 0) 
            w_count =  tm_status.get("W", 0) 
            ba_count =   tm_status.get("Ba", 0) 
            bo_count  =  tm_status.get("Bo", 0)

            r_colr_count = tm_status.get("r", 0)
            b_colr_count = tm_status.get("b", 0)

            a_count_str = "A "+ str(a_count)
            w_count_str = "W "+ str(w_count)
            ba_count_str = "Ba "+ str(ba_count)
            bo_count_str = "Bo "+ str(bo_count)

            r_colr_count_str = "R "+ str(r_colr_count)
            b_colr_count_str = "B "+ str(b_colr_count)

            cell = my_team_sheet.cell(row=13, column=col)
            cell.value = a_count_str + "," + w_count_str

            cell = my_team_sheet.cell(row=14, column=col)
            cell.value = ba_count_str + "," + bo_count_str

            not_perfects = []

            if not (PLAYER_TYPE_RULES["A"]["min"] <= a_count  and a_count <= PLAYER_TYPE_RULES["A"]["max"]):
                not_perfects.append(a_count_str) 
            
            if not (PLAYER_TYPE_RULES["W"]["min"] <= w_count  and w_count <= PLAYER_TYPE_RULES["W"]["max"]):
                not_perfects.append(w_count_str) 
            
            if not (PLAYER_TYPE_RULES["Ba"]["min"] <= ba_count  and ba_count <= PLAYER_TYPE_RULES["Ba"]["max"]):
                not_perfects.append(ba_count_str) 
            
            if not (PLAYER_TYPE_RULES["Bo"]["min"] <= bo_count  and bo_count <= PLAYER_TYPE_RULES["Bo"]["max"]):
                not_perfects.append(bo_count_str) 
            
            p_cnt = a_count + w_count + ba_count + bo_count
            
            cell = my_team_sheet.cell(row=15, column=col)
            cell.value = r_colr_count_str + "," + b_colr_count_str

            if p_cnt== 11 and len(not_perfects) == 0:
                cell = my_team_sheet.cell(row=16, column=col)
                cell.value = "Perfect"
                black_font = Font(color=BLACK) 
                cell.font = black_font
            else:
                cell = my_team_sheet.cell(row=17, column=col)
                red_font = Font(color=RED) 
                cell.font = red_font
                cell.value = "Not Perfect"

                cell = my_team_sheet.cell(row=18, column=col)
                cell.value = ",".join(not_perfects)
            
        #computation 
        
        w,ba,bo,a = 2,4,4,2

        w1,ba1,bo1,a1 = len(player_to_type["W"]),len(player_to_type["Ba"]),len(player_to_type["Bo"]),len(player_to_type["A"])

        exp_w, exp_ba, exp_bo, exp_a = w*no_team, ba*no_team, bo*no_team, a*no_team

        wgt_w, wgt_ba,wgt_bo,wgt_a = math.ceil(exp_w/w1),math.ceil(exp_ba/ba1),math.ceil(exp_bo/bo1),math.ceil(exp_a/a1)

        diff_weight = {
            "W":(wgt_w*w1)-exp_w, 
            "Ba":(wgt_ba*ba1)-exp_ba,
            "Bo":(wgt_bo*bo1)-exp_bo,
            "A":(wgt_a*a1)-exp_a
        }

        master_compute = []
        master_compute.append(list(diff_weight.keys()))
        master_compute.append([w,ba,bo,a])
        master_compute.append([w1,ba1,bo1,a1])
        master_compute.append([exp_w, exp_ba, exp_bo, exp_a])
        master_compute.append([wgt_w, wgt_ba,wgt_bo,wgt_a])
        master_compute.append(list(diff_weight.values()))

        write_range = f"D19:D45"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        
        cell = my_team_sheet["D18"]
        cell.value = "Computed"

        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                pname = my_team_sheet.cell(row=row, column=1)
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = my_team_players.get(pname.value,{}).get("weight",0)

        write_range = f"G19:J24"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        v_row, v_col = 0,0
        for row in range(min_row, max_row + 1):
            v_col = 0
            for col in range(min_col, max_col + 1):
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = master_compute[v_row][v_col]
                v_col = v_col + 1
            v_row = v_row + 1


        team_output = BytesIO()
        wb.save(team_output)
        team_output.seek(0)

        file_name = file_name
        # Step 4: Create a download button
        btn = st.download_button(
            label="Download Excel with My Team Formation",
            data=team_output,
            file_name="my_team_updated_"+file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def generate_my_teams(master_wb, file_name):
    no_team = st.session_state["input_team_generation_count"]

    is_manual_weight = st.session_state["use_manual_weight"]

    player_credit = {}
    if is_manual_weight:
        
        for key in st.session_state:
        # Check if the key belongs to your input fields
            if key.startswith("input_myteam_"):
                # Access the value of each input field
                value = st.session_state[key]
                # Process the value as needed
                pname = re.sub("input_myteam_[\d]+_", "", key)

                #print(pname, value)
                if value is None or value == "":
                    player_credit[pname] = -1
                else:
                    player_credit[pname] = eval(value)

    is_weight_inputted = False 


    wb = master_wb #load_workbook(exel_file, read_only=False)
    sheet_names = wb.sheetnames
    if my_team_sheet_name in sheet_names:
        my_team_players = {}
        my_team_sheet = wb[my_team_sheet_name]
        gph_idx = f"A19:C40"
        
        rnk = 1 
        top_11_players = []
        bottom_11_player = []
        player_to_type = {
            "A":[],
            "Ba" :[],
            "Bo":[],
            "W":[]
        }
        for row in my_team_sheet[gph_idx]:
            if row[0] is not None:
                
                if row[0].font and row[0].font.color:
                    font_color = row[0].font.color.rgb
                    if font_color  == RED:
                        color = "r"
                    else:
                        color = "b"
                else:
                    color = "b"
                                
                my_team_players[row[0].value] = {"color": color, "rank" :rnk, "type":  row[1].value, "weight":0}

                player_to_type[row[1].value].append(row[0].value)

                if rnk <=11:
                    top_11_players.append(row[0].value)
                else:
                    bottom_11_player.append(row[0].value)
                rnk = rnk + 1

        # if not is_manual_weight:
        team_comb_dict = get_team_combination(no_team,player_to_type, player_credit)
        for k,v in my_team_players.items():
            idx = player_to_type[my_team_players[k]["type"]].index(k)
            wgt = len(team_comb_dict[my_team_players[k]["type"]][idx])
            my_team_players[k]["weight"] = wgt

        #{"W":6,"Ba":4,"Bo":4,"A":2}
        play_expt_cmb_cnt = {"min":{"W":1,"Ba":3,"Bo":3,"A":1},"max":{"W":4,"Ba":6,"Bo":6,"A":4}}

        my_team = []
        my_team_player_count = {}
        
        for i in range(no_team):
            team = []
            #print(team_comb_dict)
            for k,cnt in play_expt_cmb_cnt["min"].items():
                plr_selection_cnt = cnt if cnt < len(team_comb_dict[k]) else len(team_comb_dict[k])
                for j in range(plr_selection_cnt):
                    for pname_in_list in team_comb_dict[k]:
                        if pname_in_list:
                            pname = pname_in_list[0]
                            if (pname not in team and can_add_this_player_by_color(team,my_team_players,pname)):
                                team.append(pname)
                                if my_team_player_count.get(pname, None) == None:
                                    my_team_player_count[pname] = 1
                                else: 
                                    my_team_player_count[pname] = my_team_player_count[pname] + 1

                                pname_in_list.pop(0)
                                break
                    
            my_team.append(team)

            for player, types in team_comb_dict.items():
                team_comb_dict[player] = [lst for lst in types if lst]
        

        for i in range(no_team):
            team = my_team[i]
            for k,cnt in play_expt_cmb_cnt["max"].items():
                for pname_in_list in team_comb_dict[k]:
                    if pname_in_list:
                        pname = pname_in_list[0]
                        if (len(team) < 11 and pname not in team and can_add_this_player_by_color(team,my_team_players,pname)):
                            team.append(pname)
                            if my_team_player_count.get(pname, None) == None:
                                my_team_player_count[pname] = 1
                            else: 
                                my_team_player_count[pname] = my_team_player_count[pname] + 1
                            pname_in_list.pop(0)

        # random.shuffle(my_team)

        # for t in my_team:
        #     random.shuffle(t)

        check_and_replace_players(my_team_players,my_team,team_comb_dict,play_expt_cmb_cnt)

        teams_status = []
        t_count = 1
        for update_team in my_team:
            tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
            for pname in update_team:
                pyr = my_team_players.get(pname,{})
                ptype = pyr.get("type","W")
                pcolor = pyr.get("color","b")
                if ptype is not None:
                    tems_cnt[ptype] = tems_cnt[ptype] + 1
                
                if pcolor is not None:
                    tems_cnt[pcolor] = tems_cnt[pcolor] + 1
        
            teams_status.append(tems_cnt)
            # update_team.append(t_count)
            # update_team.append(t_count)
            t_count = t_count + 1
        
        
        team_count = 1
        last_col_name = get_column_letter(no_team)

        team_number_range = f"A1:{last_col_name}1"
        min_col, min_row, max_col, max_row = range_boundaries(team_number_range)
        team_idx = 1
        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = team_idx
            team_idx = team_idx + 1

        p_count = 0
        for tm in my_team:
            p_count = p_count + len(tm)
            if len(tm) < 11:
                missing_no_of_players = 11 - len(tm)
                for i in range(missing_no_of_players):
                    tm.append("NOT ABLE TO FILL")
            

        #print(team_comb_dict)

        write_range = f"A2:{last_col_name}12"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        #print(min_col, min_row, max_col, max_row)
        for col in range(min_col, max_col + 1):
            max_pname_len = 0
            for row in range(min_row, max_row + 1):
                pname = my_team[col-1][row-2]
                cell = my_team_sheet.cell(row=row, column=col)
                if len(str(pname)) > max_pname_len:
                    max_pname_len = len(str(pname))

                if pname in bottom_11_player:
                    cell.fill = PatternFill(start_color="FBDAD7", fill_type='solid')
                f_color = my_team_players.get(pname,{}).get("color","b")
                if f_color == "r":
                    red_font = Font(color=RED) 
                    cell.font = red_font
                else:
                    black_font = Font(color=BLACK) 
                    cell.font = black_font

                cell.value = pname
            
            col_letter = get_column_letter(col)
            my_team_sheet.column_dimensions[col_letter].width = max_pname_len
        
        
        write_range = f"A13:{last_col_name}18"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        for col in range(min_col, max_col + 1):
            tm_status = teams_status[col-1]

            a_count = tm_status.get("A", 0) 
            w_count =  tm_status.get("W", 0) 
            ba_count =   tm_status.get("Ba", 0) 
            bo_count  =  tm_status.get("Bo", 0)

            r_colr_count = tm_status.get("r", 0)
            b_colr_count = tm_status.get("b", 0)

            a_count_str = "A "+ str(a_count)
            w_count_str = "W "+ str(w_count)
            ba_count_str = "Ba "+ str(ba_count)
            bo_count_str = "Bo "+ str(bo_count)

            r_colr_count_str = "R "+ str(r_colr_count)
            b_colr_count_str = "B "+ str(b_colr_count)

            cell = my_team_sheet.cell(row=13, column=col)
            cell.value = a_count_str + "," + w_count_str

            cell = my_team_sheet.cell(row=14, column=col)
            cell.value = ba_count_str + "," + bo_count_str

            not_perfects = []

            if not (PLAYER_TYPE_RULES["A"]["min"] <= a_count  and a_count <= PLAYER_TYPE_RULES["A"]["max"]):
                not_perfects.append(a_count_str) 
            
            if not (PLAYER_TYPE_RULES["W"]["min"] <= w_count  and w_count <= PLAYER_TYPE_RULES["W"]["max"]):
                not_perfects.append(w_count_str) 
            
            if not (PLAYER_TYPE_RULES["Ba"]["min"] <= ba_count  and ba_count <= PLAYER_TYPE_RULES["Ba"]["max"]):
                not_perfects.append(ba_count_str) 
            
            if not (PLAYER_TYPE_RULES["Bo"]["min"] <= bo_count  and bo_count <= PLAYER_TYPE_RULES["Bo"]["max"]):
                not_perfects.append(bo_count_str) 
            
            p_cnt = a_count + w_count + ba_count + bo_count
            
            cell = my_team_sheet.cell(row=15, column=col)
            cell.value = r_colr_count_str + "," + b_colr_count_str

            if p_cnt== 11 and len(not_perfects) == 0:
                cell = my_team_sheet.cell(row=16, column=col)
                cell.value = "Perfect"
                black_font = Font(color=BLACK) 
                cell.font = black_font
            else:
                cell = my_team_sheet.cell(row=17, column=col)
                red_font = Font(color=RED) 
                cell.font = red_font
                cell.value = "Not Perfect"

                cell = my_team_sheet.cell(row=18, column=col)
                cell.value = ",".join(not_perfects)
            
        #computation 
        
        w,ba,bo,a = 2,4,4,2

        w1,ba1,bo1,a1 = len(player_to_type["W"]),len(player_to_type["Ba"]),len(player_to_type["Bo"]),len(player_to_type["A"])

        exp_w, exp_ba, exp_bo, exp_a = w*no_team, ba*no_team, bo*no_team, a*no_team

        wgt_w, wgt_ba,wgt_bo,wgt_a = math.ceil(exp_w/w1),math.ceil(exp_ba/ba1),math.ceil(exp_bo/bo1),math.ceil(exp_a/a1)

        diff_weight = {
            "W":(wgt_w*w1)-exp_w, 
            "Ba":(wgt_ba*ba1)-exp_ba,
            "Bo":(wgt_bo*bo1)-exp_bo,
            "A":(wgt_a*a1)-exp_a
        }

        master_compute = []
        master_compute.append(list(diff_weight.keys()))
        master_compute.append([w,ba,bo,a])
        master_compute.append([w1,ba1,bo1,a1])
        master_compute.append([exp_w, exp_ba, exp_bo, exp_a])
        master_compute.append([wgt_w, wgt_ba,wgt_bo,wgt_a])
        master_compute.append(list(diff_weight.values()))

        write_range = f"D19:D40"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        
        cell = my_team_sheet["D18"]
        cell.value = "Computed"

        for col in range(min_col, max_col + 1):
            for row in range(min_row, max_row + 1):
                pname = my_team_sheet.cell(row=row, column=1)
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = my_team_players[pname.value].get("weight")

        write_range = f"G19:J24"
        min_col, min_row, max_col, max_row = range_boundaries(write_range)
        v_row, v_col = 0,0
        for row in range(min_row, max_row + 1):
            v_col = 0
            for col in range(min_col, max_col + 1):
                cell = my_team_sheet.cell(row=row, column=col)
                cell.value = master_compute[v_row][v_col]
                v_col = v_col + 1
            v_row = v_row + 1


        team_output = BytesIO()
        wb.save(team_output)
        team_output.seek(0)

        file_name = file_name
        # Step 4: Create a download button
        btn = st.download_button(
            label="Download Excel with My Team Formation",
            data=team_output,
            file_name="my_team_updated_"+file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
                
    else:
        st.warning("My Team input Sheet is missing. Please add and re-run")

def player_credit_from_excel_sheet(sheet):
    players_credits = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
        name, credit = row
        players_credits[name] = credit if credit is not None else 0
    return players_credits

def check_all_team_marked_c_and_vc(sheet):
    last_column_with_data = 0 
    for column in sheet.iter_rows(min_row=1, max_row=2):
        for cell in column:
            if cell.value is not None:
                last_column_with_data = cell.column

    last_col_name = get_column_letter(last_column_with_data)
    
    last_column_with_data = last_column_with_data - 1

    last_row_index = 12
    last_idx = f"{last_col_name}{last_row_index}"

    # # Create the range string
    gph_idx = f"B2:{last_idx}"

    min_col, min_row, max_col, max_row = range_boundaries(gph_idx)
    c_vc_missing = []
    for col in range(min_col, max_col + 1):
        _c, _vc = False, False
        for row in range(min_row, max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value == "E" or cell.value == "" or cell.value is None:
                _c = True
                _vc = True
                break
            else:
                clr = cell.fill.start_color.index
                #print(col, row, clr,C_COLOR,VC_COLOR)
                if clr in C_COLOR:
                    _c = True
                if clr in VC_COLOR:
                    _vc = True
        
        if _c and _vc:
            pass
        else:
            c_vc_missing.append(sheet.cell(row=1, column=col).value)
    
    if len(c_vc_missing) == 0:
        return True, c_vc_missing
    else:
        return False, c_vc_missing

def compute_and_download(excel_data, is_player_sheet_exists):

    player_credit = {}
    for key in st.session_state:
    # Check if the key belongs to your input fields
        if key.startswith("input_"):
            # Access the value of each input field
            value = st.session_state[key]
            # Process the value as needed
            pname = re.sub("input_[\d]+_", "", key)

            if value is None or value == "":
                player_credit[pname] = 0 
            else:
                player_credit[pname] = eval(value)
    
    #st.write(player_credit)

    credits_rows = []
    for r in excel_data:
        crow = []
        for r_val in r:
            (pname, factor), = r_val.items()
            crow.append(player_credit.get(pname,0) * factor)
        credits_rows.append(crow)
    
    
    # Calculate the number of columns
    num_columns = len(credits_rows[0])
    # Initialize a list with zeros for storing column sums
    column_sums = [0] * num_columns

    column_empty = [""] * num_columns

    # Iterate through each row and column to calculate column sums
    for row in credits_rows:
        for i in range(num_columns):
            column_sums[i] += row[i]

    # Append the column sums to the original 2D list
            
    credits_rows.append(column_empty)
    credits_rows.append(column_sums)

    file_name = teams_file.name

    wb = load_workbook(teams_file, read_only=False)
    
    sheet_names = wb.sheetnames
    sheet_name_lower = [s.lower() for s in sheet_names]

    if players_sheet_name.lower() in sheet_name_lower:
        player_sheet = wb[players_sheet_name]
        wb.remove(player_sheet)
        player_sheet = wb.create_sheet(players_sheet_name)
    else:
        player_sheet = wb.create_sheet(players_sheet_name)
    
    #append the new / updated values
    player_sheet.append(["Name", "Credit"])
    for name, credit in player_credit.items():
        player_sheet.append([name, credit])
    
    # Set the column widths for better readability
    for column in player_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        player_sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length
    

    if sheet_to_use in wb.sheetnames:
        sheet = wb[sheet_to_use]

        start_row = 60
        start_rol = 2

        for row_index, row in enumerate(credits_rows, start=start_row):  # Start=1 since Excel rows are 1-indexed
            for col_index, value in enumerate(row, start=start_rol):  # Start=1 for columns as well
                cell = sheet.cell(row=row_index, column=col_index)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)

                cell.value = value
    
    copy_sheet = wb[sheet_to_use]
    dif_in_score_teams = []

    if teams_list_sheet_name.lower() in sheet_name_lower:
        teams_sheet = wb[teams_list_sheet_name]
        idx = 2
        for v in column_sums:
            teams_sheet["B"+str(idx)] = v
            
            if teams_sheet["C"+str(idx)].value and teams_sheet["C"+str(idx)].value != "":
                diff = float(teams_sheet["C"+str(idx)].value) - float(v) 
                #print(teams_sheet["C"+str(idx)].value,v, diff)
                if diff != 0.0 :
                    dif_in_score_teams.append("T"+ str(teams_sheet["A"+str(idx)].value))
            idx = idx + 1

    else:
        teams_sheet = wb.create_sheet(teams_list_sheet_name)
        teams_sheet.append(["TeamName","Computed","Actual"])

        for team_name in TEAMS:
            teams_sheet.append(["T"+str(team_name)])


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if len(dif_in_score_teams) > 0:
        team_ids = ",".join(dif_in_score_teams)
        st.error("Teams with values mismatch between computed and actual : "+team_ids )
    else:
        st.success("All values are matching (Both Computed and Actual)")

    # Step 4: Create a download button
    btn = st.download_button(
        label="Download updated Excel file",
        data=output,
        file_name="updated_file_"+file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if process_button:
    wb = load_workbook(teams_file, read_only=False)

    is_all_okay, missing_team = check_all_team_marked_c_and_vc(wb[sheet_to_use])
    if is_all_okay:

        sheet_names = wb.sheetnames
        sheet_name_lower = [s.lower() for s in sheet_names]
        
        data = []
        if sheet_to_use in wb.sheetnames:
            sheet = wb[sheet_to_use]
            last_column_with_data = 0 
            for column in sheet.iter_rows(min_row=1, max_row=2):
                for cell in column:
                    if cell.value is not None:
                        last_column_with_data = cell.column

            last_col_name = get_column_letter(last_column_with_data)
            
            last_column_with_data = last_column_with_data - 1

            st.write("team count : ", last_column_with_data)

            last_row_index = 12
            last_idx = f"{last_col_name}{last_row_index}"

            # Create the range string
            gph_idx = f"B2:{last_idx}"

            for row in sheet[gph_idx]:
                r_values = []
                for cell in row:
                    if cell.value is not None:
                        clr = cell.fill.start_color.index
                        if clr in C_COLOR:
                            r_values.append({cell.value:2})
                        elif clr in VC_COLOR:
                            r_values.append({cell.value:1.5})
                        else:
                            r_values.append({cell.value:1})
                    else:
                         r_values.append({"":0})
                data.append(r_values)
            
            teams_name_idx = f"B1:{last_col_name}1"
            #print("teams : ", teams_name_idx)
            for row in sheet[teams_name_idx]:
                for cell in row:
                    TEAMS.append(cell.value)
            


        if players_sheet_name.lower() in sheet_name_lower:
            st.write("Existing Credit Sheet Found.")
            sheet = wb[players_sheet_name]
            player_credit = player_credit_from_excel_sheet(sheet)

            form = st.form("my_form")
            with form:
                st.header("Player Credits")
                i = 1 
                for name in player_credit:
                    input_key = f"input_{i}_{name}"
                    player_credit[name] = st.text_input(f"{i}.{name}", 
                                                                value=player_credit[name],
                                                                key=input_key)
                    i = i + 1
                submit_btn = form.form_submit_button("Submit", on_click=compute_and_download,args=(data,True) )

        else:
            st.write("No Existing Credit Sheet Found.")
            unique_players_dict = {list(x.keys())[0] for l in data for x in l}

            unique_players_list = list(unique_players_dict)

            form = st.form("my_form")

            wb.close()

            with form:
                i = 1
                input_values = {}
                for name in unique_players_list:
                    input_key = f"input_{i}_{name}"
                    
                    if (name != "E" and len(name) > 2):
                        input_values[input_key] = st.text_input(str(i) + "." + name, key=input_key)
                    else:
                        input_values[input_key] = st.text_input(str(i) + "." + name, key=input_key, value= 0)
                    i = i + 1

                submit_btn = form.form_submit_button("Submit", on_click=compute_and_download,args=(data,False) )
        
    else:
        st.write("All Teams Must Have C and VC, Following Teams are having issue. ")
        st.write(missing_team)


            
                
if my_team_formation:


    form = st.form("my_form")
    with form:
        player_cat_mapping = {
            "ar" : "A",
            "ba" : "Ba",
            "bo" : "Bo",
            "w" : "W",
            "a" : "A"
            
        }
         
        wb = load_workbook(teams_file)
        sheet_names = wb.sheetnames

        min_value_match = re.search(r'\b(\d+)T\b', teams_file.name)

        # Extracting the value
        if min_value_match:
            extracted_min_value = min_value_match.group(1)
        else:
            extracted_min_value = 3

        expected_team_count = st.number_input("Enter no of team required.", min_value=int(extracted_min_value),
                                               max_value=5000, 
                                               step=1,
                                                 format='%d', key="input_team_generation_count")
        
        sel_player_weight = {}

        if sel_value_sheet_name in sheet_names:
            sel_value_sheet = wb[sel_value_sheet_name]
            for row in sel_value_sheet:
                if row[0].value is None:
                    break
                if row[0].value.lower() != "name":
                    #print(row[0].value, row[1].value)
                    if row[1].value:
                        sel_player_weight[row[0].value.strip()] = eval(str(row[1].value))
                    else:
                        sel_player_weight[row[0].value.strip()] = 0

        if my_team_sheet_name not in sheet_names:
            my_team_payers = {}
            new_my_teams_sheet = wb.create_sheet(my_team_sheet_name)

            cvc_sheet = wb["C VC"]
            for row in cvc_sheet:
                if row[0].value is None:
                    break
                fill = row[0].fill
                #print(row[0].value.lower())
                is_player_to_be_considered = False

                if row[0].value.lower() == "name":
                    is_player_to_be_considered = False
                    pass
                
                elif not is_retro_game:
                    if (fill.patternType == 'solid' and fill.fgColor.rgb == "FFFFFF00"):
                        is_player_to_be_considered = False
                        pass
                    else:
                        is_player_to_be_considered = True
                else:
                    if (fill.patternType == 'solid' and fill.fgColor.rgb == "FFFFFF00"):
                        if row[1].value is not None and len(row[0].value) > 2 and row[1].value.lower() == "playing":
                            is_player_to_be_considered = True
                    else:
                        if row[1].value is not None and len(row[0].value) > 2 and row[1].value.lower() in ["not playing","non playing"]:
                            is_player_to_be_considered = False
                        else:
                            is_player_to_be_considered = True
                
                if is_player_to_be_considered:
                    if len(row[0].value) > 3:
                        ptype_grp = match = re.search(r'\((.*?)\)', row[0].value)
                        ptype = ""
                        if ptype_grp:
                            ptype = ptype_grp.group(1)
                        
                        if row[0].font and row[0].font.color:
                            font_color = row[0].font.color.rgb
                            if font_color  == RED:
                                color = "r"
                            else:
                                color = "b"
                        else:
                            color = "b"

                        name_grp = re.search(r'^(.*?)\s*\(', row[0].value)

                        name = ""
                        if name_grp:
                            name = name_grp.group(1)

                        my_team_payers[name.strip()] = {"color" : color, "type": player_cat_mapping.get(ptype.lower(),"W")}


            retro_new_weight = {}
            if is_retro_game:
                ws = wb[sheet_to_use]
                # Element you're searching for
                search_element =  list(my_team_payers.keys())[0]

                # Iterate over rows starting from row 25 and columns up to 34
                restricted_locations = []

                # Iterate over all rows starting from row 25 to find occurrences of "J Charles" in columns up to 34
                for row_idx, row in enumerate(ws.iter_rows(min_row=25, max_col=34, values_only=True), start=25):
                    # Track columns where "J Charles" is found in the current row (restricted to columns 1 through 34)
                    current_row_occurrences = [col_idx + 1 for col_idx, cell in enumerate(row[:34]) 
                                            if cell and search_element.lower() in str(cell).lower()]
                    
                    # If "J Charles" appears at least twice in the same row, record the row and column numbers
                    if len(current_row_occurrences) >= 2:
                        restricted_locations.append((current_row_occurrences, row_idx))
                        break  

                identified_columns = restricted_locations[0][0]  # Extracting column indices from the first found location

                
                for row_idx, row in enumerate(ws.iter_rows(min_row=24, max_row=80, min_col =identified_columns[0],
                                                    max_col=identified_columns[1], values_only=True), start=24):
                    
                    if row[0] and row[2] and row[0] == row[2]:
                        retro_new_weight[row[0].strip()] =  row[1] if row[1] else 0

            #print(retro_new_weight)
            row = 19
            for k,v in my_team_payers.items():
                if v["color"] == "r":
                    red_font = Font(color=RED) 
                    new_my_teams_sheet.cell(row=row, column=1).font = red_font
                else:
                    black_font = Font(color=BLACK) 
                    new_my_teams_sheet.cell(row=row, column=1).font = black_font
                new_my_teams_sheet.cell(row=row, column=1).value =k
                if len(v["type"])>1:
                    valid_ptype = v["type"][0].upper()+v["type"][1].lower()
                else:
                    valid_ptype = v["type"][0].upper()
                new_my_teams_sheet.cell(row=row, column=2).value = valid_ptype
                
                if not is_retro_game:
                    new_my_teams_sheet.cell(row=row, column=3).value = int(sel_player_weight.get(k,0))
                else:
                    new_my_teams_sheet.cell(row=row, column=3).value = int(retro_new_weight.get(k,0))

                row = row + 1

            
        if use_user_weight:
            st.header("Player Weight")
            my_team_sheet = wb[my_team_sheet_name]
            gph_idx = f"A19:C45"
            
            i = 0
            for row in my_team_sheet[gph_idx]:
                if row[0].value is not None:
                    name = row[0].value
                    user_wgt = 0
                    if row[2].value is not None:
                        user_wgt = row[2].value
                    elif sel_player_weight.get(name,None) is not None:
                        user_wgt = sel_player_weight.get(name.strip())
                    else:
                        user_wgt = 0 
                    vlu =user_wgt #row[2].value if row[2].value is not None else 0  
                    
                    input_key = f"input_myteam_{i}_{name}"
                    st.text_input(f"{i}.{name}",key=input_key, value= vlu)
                    i = i + 1

        submit_btn = form.form_submit_button("Generate Team",
                                              on_click=manual_generate_my_teams,
                                              args=(wb,teams_file.name)) #teams_file
        

            



##################My TEAM Code###################
        ####Depricated
        # my_team = []
        # print(no_team)
        # for team_id in range(no_team):
        #     my_team.append(list(top_11_players))
        

        # pvt_point = int(no_team / 2)

        # s_idx, m_idx, l_idx = 0, pvt_point, no_team
        # p_idx, itr  = 0, 1
        # rest_player_idx_at = no_team if no_team > len(bottom_11_player) else len(bottom_11_player) 
        # for i in range(10,-1, -1):

        #     if i % 2 == 0:
        #         s_idx, m_idx = 0, pvt_point
        #     else:
        #         s_idx = m_idx
        #         m_idx = l_idx
        
        #     for j in range(s_idx,m_idx):
                
        #         my_team[j][i] = bottom_11_player[p_idx%11]
        #         p_idx = p_idx + 1
            
        #         if itr % rest_player_idx_at == 0:
        #             bottom_11_player = bottom_11_player[1:] + [bottom_11_player[0]]

        #         itr = itr + 1

        # team_count = 1
        # last_col_name = get_column_letter(no_team)
        # write_range = f"A1:{last_col_name}12"
        # min_col, min_row, max_col, max_row = range_boundaries(write_range)
        

        # teams_status = []
        # t_count = 1
        # #print(my_team_players)
        # for update_team in my_team:
        #     tems_cnt = {"W":0, "Ba": 0, "A": 0, "Bo": 0,"r":0,"b":0}
        #     for pname in update_team:
        #         pyr = my_team_players.get(pname,{})
        #         ptype = pyr.get("type","W")
        #         pcolor = pyr.get("color","b")
        #         if ptype is not None:
        #             tems_cnt[ptype] = tems_cnt[ptype] + 1
                
        #         if pcolor is not None:
        #             tems_cnt[pcolor] = tems_cnt[pcolor] + 1
        
        #     teams_status.append(tems_cnt)
        #     update_team.insert(0,t_count)
        #     t_count = t_count + 1
                
        # # print(min_col, min_row, max_col, max_row,len(my_team), len(my_team[0]))
        # for col in range(min_col, max_col + 1):
        #     for row in range(min_row, max_row + 1):
        #         pname = my_team[col-1][row-1]
        #         cell = my_team_sheet.cell(row=row, column=col)
                
        #         if pname in bottom_11_player:
        #             cell.fill = PatternFill(start_color="FBDAD7", fill_type='solid')
        #         f_color = my_team_players.get(pname,{}).get("color","b")
        #         if f_color == "r":
        #             red_font = Font(color=RED) 
        #             cell.font = red_font
        #         else:
        #             black_font = Font(color=BLACK) 
        #             cell.font = black_font

        #         cell.value = pname
            
        # write_range = f"A13:{last_col_name}18"
        # min_col, min_row, max_col, max_row = range_boundaries(write_range)
        # for col in range(min_col, max_col + 1):
        #     tm_status = teams_status[col-1]

        #     a_count = tm_status.get("A", 0) 
        #     w_count =  tm_status.get("W", 0) 
        #     ba_count =   tm_status.get("Ba", 0) 
        #     bo_count  =  tm_status.get("Bo", 0)

        #     r_colr_count = tm_status.get("r", 0)
        #     b_colr_count = tm_status.get("b", 0)

        #     a_count_str = "A "+ str(a_count)
        #     w_count_str = "W "+ str(w_count)
        #     ba_count_str = "Ba "+ str(ba_count)
        #     bo_count_str = "Bo "+ str(bo_count)

        #     r_colr_count_str = "Red "+ str(r_colr_count)
        #     b_colr_count_str = "Black "+ str(b_colr_count)

        #     cell = my_team_sheet.cell(row=13, column=col)
        #     cell.value = a_count_str + "," + w_count_str

        #     cell = my_team_sheet.cell(row=14, column=col)
        #     cell.value = ba_count_str + "," + bo_count_str

        #     not_perfects = []

        #     if not (PLAYER_TYPE_RULES["A"]["min"] <= a_count  and a_count <= PLAYER_TYPE_RULES["A"]["max"]):
        #         not_perfects.append(a_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["W"]["min"] <= w_count  and w_count <= PLAYER_TYPE_RULES["W"]["max"]):
        #         not_perfects.append(w_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["Ba"]["min"] <= ba_count  and ba_count <= PLAYER_TYPE_RULES["Ba"]["max"]):
        #         not_perfects.append(ba_count_str) 
            
        #     if not (PLAYER_TYPE_RULES["Bo"]["min"] <= bo_count  and bo_count <= PLAYER_TYPE_RULES["Bo"]["max"]):
        #         not_perfects.append(bo_count_str) 
                
        #     if len(not_perfects) == 0:
        #         cell = my_team_sheet.cell(row=15, column=col)
        #         cell.value = "Perfect"
        #         black_font = Font(color=BLACK) 
        #         cell.font = black_font
        #     else:
        #         cell = my_team_sheet.cell(row=15, column=col)
        #         red_font = Font(color=RED) 
        #         cell.font = red_font
        #         cell.value = "Not Perfect"

        #         cell = my_team_sheet.cell(row=16, column=col)
        #         cell.value = ",".join(not_perfects)
            
        #     cell = my_team_sheet.cell(row=17, column=col)
        #     cell.value = r_colr_count_str + "," + b_colr_count_str


        # team_output = BytesIO()
        # wb.save(team_output)
        # team_output.seek(0)

        # file_name = exel_file.name
        # # Step 4: Create a download button
        # btn = st.download_button(
        #     label="Download Excel with My Team Formation",
        #     data=team_output,
        #     file_name="my_team_updated_"+file_name,
        #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # )
#### OLD depricated END

#### Depricated 
        
    #     def get_team_combination(no_team,player_to_type):

    # output = {
    #     "W" :[],
    #     "Ba":[],
    #     "Bo":[],
    #     "A":[]
    # }

    # total_possible_players = no_team * 12 
    # w,ba,bo,a = 2,4,4,2

    # w1,ba1,bo1,a1 = len(player_to_type["W"]),len(player_to_type["Ba"]),len(player_to_type["Bo"]),len(player_to_type["A"])

    # exp_w, exp_ba, exp_bo, exp_a = w*no_team, ba*no_team, bo*no_team, a*no_team

    # wgt_w, wgt_ba,wgt_bo,wgt_a = math.ceil(exp_w/w1),math.ceil(exp_ba/ba1),math.ceil(exp_bo/bo1),math.ceil(exp_a/a1)

    # diff_weight = {
    #     "W":(wgt_w*w1)-exp_w, 
    #     "Ba":(wgt_ba*ba1)-exp_ba,
    #     "Bo":(wgt_bo*bo1)-exp_bo,
    #     "A":(wgt_a*a1)-exp_a
    # }

    # p_weights = {
    #     "W":[wgt_w]*w1,
    #     "Ba":[wgt_ba]*ba1,
    #     "Bo":[wgt_bo]*bo1,
    #     "A":[wgt_a]*a1
    # }

     
    # for k,v in diff_weight.items():
    #     v = diff_weight[k]
    #     if v > 0:
    #         for i in range(-v,0):
    #             new_va = p_weights[k][i] - 1
    #             p_weights[k][i]  = new_va

    # p_weights_to_dict = p_weights

    # for k,v in p_weights_to_dict.items():
    #     for idx in range(len(v)):
    #         output[k].append([player_to_type[k][idx]]*v[idx])

    # return output
